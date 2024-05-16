from django.shortcuts import render,redirect,get_object_or_404
from .models import *
from login.models import *
from seller.models import *
from django.contrib.auth.decorators import login_required
from .forms import ExcelUploadForm

from datetime import datetime
from django.core.files.base import ContentFile
from tablib import Dataset
from customer.models import *
import random
import string
from django.core.mail import send_mail
import sweetify
from .forms import *
from django.db.models import Sum
# Create your views here.

def bootstrap(request):
    return render(request, "admin_base.html")

@login_required(login_url='/admin_login/')

def admin_add_product(request):
    if request.method == 'POST':
        characters_upper = string.ascii_uppercase 
        characters_digits = string.digits

        if 'myfile' in request.FILES:
            dataset = Dataset()
            new_persons = request.FILES['myfile']
            imported_data = dataset.load(new_persons.read(), format='xlsx')

            print("Imported Data:", imported_data)  # Print to check the structure of imported_data

            for data in imported_data:
                first_three_chars = 'PROD'
                remaining_digits = ''.join(random.choices(characters_digits, k=6))
                prod_id = f'{first_three_chars}{remaining_digits}'

                SKUId = data[0]
                discount_percentage=data[25]


                # Determine discount calculation based on discount_type
                discount_type = data[11]  # Assuming discount_type is at index 11 in your data

                if discount_type == 'rupees':
                    discount_price = float(data[12])
                    original_price = float(data[10])
                    selling_price = original_price - discount_price
                    discount_percentage = (discount_price / original_price) * 100
                    
                elif discount_type == 'percentage':
                    discount_percentage = float(data[12])  # This is already the percentage
                    original_price = float(data[10])
                    discount_price = (discount_percentage / 100) * original_price
                    selling_price = original_price - discount_price
                else:
                    # Handle other cases or set a default value
                    discount_price=0.0
                    discount_percentage = 0.0
                    selling_price = float(data[10])
                print('discount price',discount_percentage)

                # if discount_type == 'Rupees':
                #     # Calculate discounted price in case of Rupees
                #     discount_price = float(data[12])
                #     selling_price = float(data[10]) - discount_price
                # elif discount_type == 'Percent':
                #     # Convert percentage discount to actual amount
                #     discount_price = float(data[12])
                #     dis_price = (float(data[12]) / 100) * float(data[10])
                #     selling_price = float(data[10]) - dis_price
                # else:
                #     # Handle other cases or set a default value
                #     discount_price = 0.0
                #     selling_price = float(data[10])

                new_book = product_seller(
                    user=request.user,
                    SKUId=SKUId,
                    brand=data[1],
                    product_id=data[2],
                    size=data[3],
                    colour=data[4],
                    title=data[5],
                    description=data[6],
                    product_type=data[7],
                    category=data[8],
                    sub_category=data[9],
                    actual_price=float(data[10]),
                    discount_type=discount_type,
                    discount_price=discount_price,
                    selling_price=selling_price,
                    quantity=int(data[13]),  # Assuming quantity is an integer field
                    gender=data[14],
                    product_instruction=data[15],
                    manufacturer=data[16],
                    model_name=data[17],
                    model_number=data[18],
                    age_range_description=data[19],
                    bullet_point=data[20],
                    special_feature=data[21],
                    material_type=data[22],
                    manufacturer_contact=data[23],
                    product_story=data[24],
                    discount_percentage=discount_percentage,
                    prod_id=prod_id,
                )

                # Save the new book entry
                new_book.save()
            return redirect('/admin_all_products/')

    return render(request, 'admin/admin_add_products.html')



from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment  


def excel_export_view(request):
    # Create a new Excel workbook and add a worksheet to it
    wb = openpyxl.Workbook()
    ws = wb.active

    # Add headings to the worksheet
    headings = ["SKUId", "brand", "product_id", "size", "colour", "title", "description", "product_type", "category", "sub_category", "discount_type", "discount_price", "actual_price", "quantity", "gender",  "product_instruction", "manufacturer", "model_name", "model_number", "age_range_description", "bullet_point", "special_feature", "material_type", "manufacturer_contact", "product_story"]

    ws.append(headings)
    font = Font(size=14, bold=True, color='000000')  # Adjust size and color as needed
    for cell in ws[1]:
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align horizontally and vertically

    column_widths = [10, 15, 25, 15, 20, 25, 35, 20,20,20,20,20,20,20,20,20,20,20,20,20,20,20,20,20]  # Adjust the widths as needed
    for i, width in enumerate(column_widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = width    

    # Create a response with Excel MIME type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Sample_product_import_sheet.xlsx'

    # Save the workbook to the response
    wb.save(response)
    return response



    


def add_blogs(request):
    if request.method == "POST":
        title = request.POST.get('title')
        heading1 = request.POST.get('heading1')
        heading2 = request.POST.get('heading2')
        heading3 = request.POST.get('heading3')
        heading4 = request.POST.get('heading4')
        heading5 = request.POST.get('heading5')
        heading6 = request.POST.get('heading6')
        heading7 = request.POST.get('heading7')
        heading8 = request.POST.get('heading8')
        heading9 = request.POST.get('heading9')
        heading10 = request.POST.get('heading10')

        content1 = request.POST.get('content1')
        content2 = request.POST.get('content2')
        content3 = request.POST.get('content3')
        content4 = request.POST.get('content4')
        content5 = request.POST.get('content5')
        content6 = request.POST.get('content6')
        content7 = request.POST.get('content7')
        content8 = request.POST.get('content8')
        content9 = request.POST.get('content9')
        content10 = request.POST.get('content10')
        image1 = request.FILES.get('image1')
        image2 = request.FILES.get('image2')
        image3 = request.FILES.get('image3')
        print('imageeeeee',image1)
        
        BlogPost(title=title,image1=image1,image2=image2,image3=image3,content1=content1,content2=content2,
                 content3=content3,content4=content4,content5=content5,content6=content6,
                 content7=content7,content8=content8,content9=content9,content10=content10,
                 heading1=heading1,heading2=heading2,heading3=heading3,heading4=heading4,
                 heading5=heading5,heading6=heading6,heading7=heading7,heading8=heading8,
                 heading9=heading9,heading10=heading10).save()
        
        print('555555555555555555555555555555',BlogPost.title)
    return render(request,'blogs/add_blogs.html')

def view_blogs(request):
    blogs = BlogPost.objects.all()
    for i in blogs:
        print('bbbbbbbbbbbbbbbbbbbb',i.image1)
    context ={
        'blogs':blogs
    }
    return render(request,'blogs/view_blogs.html',context)


def blog_details(request,id):
    blog = BlogPost.objects.get(pk=id)
    return render(request, 'blogs/blog_details.html',{'blog':blog})

from django.utils import timezone
def edit_blog(request,id):
    blog = BlogPost.objects.get(pk=id)

    if request.method == 'POST':
        
        blog.title = request.POST.get('title', '')
        blog.heading1 = request.POST.get('heading1', '')
        blog.heading2 = request.POST.get('heading2', '')
        blog.heading3 = request.POST.get('heading3', '')
        blog.heading4 = request.POST.get('heading4', '')
        blog.heading5 = request.POST.get('heading5', '')
        blog.heading6 = request.POST.get('heading6', '')
        blog.heading7 = request.POST.get('heading7', '')
        blog.heading8 = request.POST.get('heading8', '')
        blog.heading9 = request.POST.get('heading9', '')
        blog.heading10 = request.POST.get('heading10', '')
        blog.content1 = request.POST.get('content1', '')
        blog.content2 = request.POST.get('content2', '')
        blog.content3 = request.POST.get('content3', '')
        blog.content4 = request.POST.get('content4', '')
        blog.content5 = request.POST.get('content5', '')
        blog.content6 = request.POST.get('content6', '')
        blog.content7 = request.POST.get('content7', '')
        blog.content8 = request.POST.get('content8', '')
        blog.content9 = request.POST.get('content9', '')
        blog.content10 = request.POST.get('content10', '')
        
        blog.image1 = request.FILES.get('image1', blog.image1)
        blog.image2 = request.FILES.get('image2', blog.image2)
        blog.image3 = request.FILES.get('image2', blog.image3)


        blog.date = timezone.now().strftime('%Y-%m-%d %H:%M:%S')

        
        blog.save()

        return redirect('/view_blogs/', id=blog.id)

    return render(request, 'blogs/edit_blogs.html', {'blog': blog})

def delete_blog(request,id):
   blog = BlogPost.objects.filter(pk=id)
    
   if request.method == 'POST':
        blog.delete()
        return redirect('/view_blogs/')



def seller_register_list(request):
    sellers = VrUser.objects.filter(is_seller=False,is_admin=False,is_customer=False).order_by('-id')  # Retrieve all Seller objects from the database
    data = seller_user.objects.filter(is_seller=False).order_by('-id')
    # print("yyyyyyyyyyyyyyy",sellers)
    context = {
        'sellers': sellers,  # Pass the sellers queryset as context
        'data':data,
    }
    return render(request,'admin/seller_register_list.html',context)


def seller_profile_view(request,id):
    data1=seller_user.objects.get(id=id)
    print('ssss',id)
   
    if request.method == 'POST':
        # comment=request.POST.get("comment")
        # print("11111111111111111111111111",comment)
        # data1.comment=comment
        # print("11111111111111111111111111",data1.comment)
        data1.save()
        
        if request.POST.get("is_seller") == "approval_button":
            data1.is_seller = True
            data1.seller_approved= True
            data1.seller_reject= False
            data1.save()
            if request.method == 'POST':
                 
                subject = "Seller Registration Confirmation"
                message = f"Hello {data1.first_name} {data1.last_name},\n\nThank you for registering as a seller on our website. Your account has been successfully created.\n\n"
                 # Replace [YourGeneratedPassword] with the actual generated password
                login_url = "http://vishwaratna.in/seller_login/"  # Replace 'seller_login' with the actual name of your seller login URL pattern
                message += f"To log in, please visit: {login_url}\n\n"
                message += "Best regards,\nVR Team"
                from_email = "myvishwaratna@gmail.com"  # Replace this with your desired 'frossm' email address
                recipient_list = [data1.email]
                send_mail(subject, message, from_email, recipient_list, fail_silently=False)

        if request.POST.get("seller_reject") == "Save":
            
            data1.seller_reject = True
            data1.is_seller = False
            data1.seller_approved= False
            data1.save()
            print("tttttttttttttttttttttttttttttttt",data1)
            if request.method == 'POST':
                comment=request.POST.get("comment")
                print("11111111111111111111111111",comment)
                data1.comment=comment
                print("11111111111111111111111111",data1.comment)
                subject = "Why Admin Is Rejected"
                message = f"Hello {data1.first_name} {data1.last_name},\n\nThank you for registering as a seller on our website\n\n"
                message += f"Reason For Rejected\n"
                message += f"{data1.comment}\n\n"
                
                message += "Best regards,\nVR Team"
                from_email = "myvishwaratna@gmail.com"  # Replace this with your desired 'from' email address
                recipient_list = [data1.email]
                send_mail(subject, message, from_email, recipient_list, fail_silently=False)
            
            
           
        
        
            # Redirect to another page or do something else after saving
        return redirect('/seller_register_list/')
    context = {'data1': data1}
    return render(request, 'admin/seller_profile_view.html', context)







def seller_approved_list(request):
    data1=seller_user.objects.filter(seller_approved=True).order_by('-id')
    context={
        "data1":data1,
    }
    return render(request, 'admin/seller_approved_list.html', context)

def seller_reject_list(request):
    data1 = seller_user.objects.filter(is_seller=False,seller_reject=False).order_by('-id')
    context={
        "data1":data1,
    }
    return render(request, 'admin/seller_reject_list.html', context)


from django.db.models import Q
def approve_seller_product(request):
    products = product_seller.objects.filter(~Q(product_reject=True) & ~Q(product_approved=True))
    return render(request,'product/approve_seller_product.html',{'products':products})

def admin_approve_product(request,id):
    product = product_seller.objects.get(id=id)
    product.product_reject = False
    product.product_approved = True
    product.save()
    return redirect('/approve_seller_product/')


def admin_reject_product(request,id):
    product = product_seller.objects.get(id=id)
    product.product_reject = True
    product.product_approved = False
    product.save()
    return redirect('/approve_seller_product/')

def approved_product(request):
    products = product_seller.objects.filter(product_approved=True)
    return render(request,'product/approved_products.html',{'products':products})

def rejected_product(request):
    products = product_seller.objects.filter(product_reject=True)
    return render(request,'product/rejected_product.html',{'products':products})

def product_details(request,id):
    product_details = product_seller.objects.filter(id=id)
    return render(request,'product/product_details.html',{'product_details':product_details})



def sellers_details(request):
    all_sellers = seller_user.objects.filter(is_seller=True)

    return render(request,'admin/sellers_details.html',{'all_sellers':all_sellers})

def view_products(request,user_id):
    user = get_object_or_404(VrUser, pk=user_id)
    products = product_seller.objects.filter(user=user)
    return render(request, 'admin/view_products.html', {'products': products})



def query_list(request):
    
    customer_queries = CustomerQuery.objects.filter()
    
    
    return render(request, 'admin/queries_list.html', {'customers': customer_queries})

from django.shortcuts import render, get_object_or_404
from django.core.mail import send_mail
from django.conf import settings

def customer_queries(request,id):

    customer_query = get_object_or_404(CustomerQuery, id=id)

    if request.method == 'POST':
        query_id = request.POST.get('query_id')
        answer_text = request.POST.get('answer')

        customer_query = CustomerQuery.objects.get(id=query_id)

        customer_query.answer = answer_text
        customer_query.save()


        send_mail(
            f"Your Query Answered: {customer_query.subject}",
            f"Your Query:\n{customer_query.message}\n\nAnswer:\n{customer_query.answer}",
            settings.DEFAULT_FROM_EMAIL,
            [customer_query.customer.email], 
            fail_silently=False,
        )

        customer_query.delete()

        return redirect('/admin_dashboard/')


    return render(request, 'admin/customer_queries.html',{'customer_query':customer_query})


def admin_all_products(request):
    user = request.user
    data = product_seller.objects.filter(user=user).order_by('-id')
    return render(request,'admin/admin_all_products.html',{'data':data})



# def upload_product_image(request,id):
    
#     seller = product_seller.objects.get(id=id)
    
        
#     if request.method == 'POST':
#             # Update the existing seller with the new data
#             seller.title = request.POST.get('title')
#             seller.seller_sku = request.POST.get('seller_sku')
#             seller.brand_name = request.POST.get('brand_name')
#             seller.color = request.POST.get('color')
            
            
#             seller.product_image = request.FILES.get('product_image')  # Handle uploaded image file
#             seller.back_image = request.FILES.get('back_image')
#             seller.left_image = request.FILES.get('left_image')
#             seller.right_image = request.FILES.get('right_image')
#             seller.right_image = seller.right_image 
#             seller.product_image=seller.product_image
#             seller.back_image=seller.back_image
#             seller.left_image=seller.left_image
#             # Save the updated seller entry
#             seller.save()
#             print("77777777777777",seller,seller.product_image)
            
           
        
    
#     context = {
#         'seller': seller,
#     }
#     return render(request, 'product/edit_seller_import.html', context)


def edit_admin_import(request,id):
    
    seller = product_seller.objects.get(id=id)
    
        
    if request.method == 'POST':
            # Update the existing seller with the new data
            # seller.title = request.POST.get('title')
            seller.SKUId = request.POST.get('SKUId')
            # seller.model_number=request.POST.get("model_number")
            
            # Check if new images are uploaded and update only if they are
            if request.FILES.get('product_image'):
                seller.product_image = request.FILES.get('product_image')
            if request.FILES.get('back_image'):
                seller.back_image = request.FILES.get('back_image')
            if request.FILES.get('left_image'):
                seller.left_image = request.FILES.get('left_image')
            if request.FILES.get('right_image'):
                seller.right_image = request.FILES.get('right_image')
            
            seller.save()
            return redirect('/admin_all_products/')
       
    
    context = {
        'seller': seller,   
    }
    return render(request, 'product/edit_admin_import.html', context)

def admin_all_products_update(request):
    user = request.user
    data = product_seller.objects.filter(user=user).order_by('-id')
    return render(request,'admin/admin_all_products_update.html',{'data':data})


def edit_product_details(request,id):
    data=product_seller.objects.get(id=id)
       
    if request.method == 'POST':
            data.title = request.POST.get('title')
            data.SKUId = request.POST.get('SKUId')
            data.brand = request.POST.get('brand')
            data.colour = request.POST.get('colour')
            data.category=request.POST.get("category")
            data.product_type=request.POST.get("product_type")
            data.sub_category=request.POST.get("sub_category")
            data.product_id=request.POST.get("product_id")
            data.quantity=request.POST.get("quantity")
            data.description=request.POST.get("description")
            data.product_instruction=request.POST.get("product_instruction")
            data.bullet_point=request.POST.get("bullet_point")
            data.product_story=request.POST.get("product_story")
            data.special_feature=request.POST.get("special_feature")
            data.product_material=request.POST.get("product_material")
            
            data.save()
            return redirect('/admin_all_products_update/')
        
    context={
            'seller':data
        }
    return render(request, 'admin/edit_product_details.html', context)



def view_invoice_no(request):
    all_invoice = OrderPlaced.objects.all()

    return render(request,'admin/view_invoice.html',{'all_invoice':all_invoice})




def approve_commissions(request):
    approved_commissions = Commission.objects.filter(status='Approved')
    return render(request, 'admin/approve_commissions.html', {'approved_commissions': approved_commissions})

def reject_commissions(request):
    rejected_commissions = Commission.objects.filter(status='Rejected')
    return render(request, 'admin/reject_commissions.html', {'rejected_commissions': rejected_commissions})

def move_to_pending(request, commission_id):
    commission = Commission.objects.get(pk=commission_id)
    commission.status = 'Pending'
    commission.save()
    return redirect('manage_commissions')

def manage_commissions(request):
    pending_commissions = Commission.objects.filter(status='Pending')
    rejected_commissions = Commission.objects.filter(status='Rejected')
    
    all_commissions = Commission.objects.all()
    for i in all_commissions:
        print('all_commissions',i.status)
    
    return render(request, 'admin/manage_commissions.html', {
        'all_commissions': all_commissions,
        'pending_commissions': pending_commissions,
        'rejected_commissions': rejected_commissions,
        'approve_view': True
    })


def approve_commission(request, commission_id):
    commission = Commission.objects.get(pk=commission_id)
    commission.status = 'Approved'
    commission.approved_by = request.user
    commission.save()
    return redirect('manage_commissions')

def reject_commission(request, commission_id):
    commission = Commission.objects.get(pk=commission_id)
    commission.status = 'Rejected'
    commission.rejected_by = request.user
    commission.save()
    return redirect('manage_commissions')



def preview(request, product_id):
    product = get_object_or_404(product_seller, id=product_id)
    context = {'product': product}
    return render(request, 'admin/preview.html', context)




from django.http import JsonResponse

def delete_admin_product(request, prod_id):
   
    product=product_seller.objects.get(prod_id=prod_id)

    product.delete()
    return redirect('/admin_all_products/')
    
# def delete_admin_product(request, product_id):
#     product = get_object_or_404(product_seller, id=product_id)
#     if request.method == 'POST':
#         product.delete()
#         # Return JSON response with the ID of the deleted product
#         return JsonResponse({'message': 'Product deleted successfully.', 'deleted_product_id': product_id})
#     return JsonResponse({'error': 'Method not allowed'}, status=405)


def all_orders(request):
    data = OrderPlaced.objects.all().order_by('-id')
    return render(request,'admin/all_orders.html',{'data':data})


def order_status(request):
    all_orders = OrderPlaced.objects.all().order_by('-id')
    return render(request,'admin/order_status.html',{'all_orders':all_orders})



STATUS_CHOICE = (
    ('Accepted', 'Accepted'),
    ('Packed', 'packed'),
    ('On The Way', 'On The Way'),
    ('Delivered', 'Delivered'),
    ('Cancel', 'Cancel'),
)



from django.utils import timezone
def view_order(request, id):
    order_details = get_object_or_404(OrderPlaced, id=id)
    tracking_details = TrackingDetail.objects.filter(order=order_details).order_by('-date_time')

    if request.method == 'POST':
        update_location = request.POST.get('location')
        status = request.POST.get('status')
        cod = request.POST.get('cod')

        if update_location:
            if not order_details.location1:
                order_details.location1 = update_location
            elif not order_details.location2:
                order_details.location2 = update_location
            elif not order_details.location3:
                order_details.location3 = update_location
            else:
                # Handle additional locations or raise an error if necessary
                pass

        order_details.update_location = update_location
        order_details.status = status
        order_details.remark = status
        order_details.cod_amount = cod

        if status == 'Delivered':
            # If status is Delivered, update cod_amount with total_cost
            order_details.cod_amount = order_details.total_cost
            order_details.Delivered=True

            order_details.delivered_date = timezone.now()

        order_details.save()

        # Create a new TrackingDetail object to store the location update
        new_tracking_detail = TrackingDetail(order=order_details, location=update_location, date_time=timezone.now(), remark=status)
        new_tracking_detail.save()
        
    return render(request, 'admin/view_order.html', {'order_details': order_details, 'tracking_details': tracking_details, 'STATUS_CHOICE': STATUS_CHOICE})



def update_status(request):
    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        print('8888888888888888888',order_id)
        new_status = request.POST.get('status')
        try:
            order = OrderPlaced.objects.get(order_id=order_id)
            order.status = new_status
            order.save()
        except OrderPlaced.DoesNotExist:
            
            pass
    return redirect('/order_status/')  





# def all_cancel_orders(request):
#     data = OrderPlaced.objects.filter(canceled=True)
#     return render(request,'admin/all_cancel_orders.html',{'data':data})

def all_cancel_orders(request,month=None,day=None):
    # Fetch all canceled orders
   canceled_orders = CanceledOrder.objects.filter(canceled=True).order_by('-canceled_at')
    
   return render(request, 'admin/all_cancel_orders.html', {'data': canceled_orders})


def show_return_orders(request): 
    order = OrderPlaced.objects.filter(returned=True)
    return render(request, "admin/show_return_orders.html", {'order': order})



def admin_availability_report(request):
    seller_id = request.user.id
    products = product_seller.objects.filter(user_id=seller_id).order_by('-id')
    
    low_quantity_products = []
    
    for product in products:
        try:
            quantity = int(product.quantity)
            if quantity < 15:
                low_quantity_products.append(product)
        except ValueError:
            # Handle cases where quantity is not a valid integer
            pass
    
    if low_quantity_products:
        send_low_quantity_email(request.user.email, low_quantity_products)
    
    return render(request, 'admin/A_availability_report.html', {'products': products})



def send_low_quantity_email(recipient_email, low_quantity_products):
    subject = 'Low Quantity Alert'
    message = f'Dear Seller,\n\nThe following products have low quantities:\n\n'
    
    for product in low_quantity_products:
        message += f'{product.title} - Quantity: {product.quantity}\n'
    
    message += '\nPlease take necessary actions to replenish the stock.\n\nRegards,\nVishwa Ratna'
    
    send_mail(subject, message, 'myvishwaratna@gmail.com', [recipient_email])


def admin_payments(request):
    data2=OrderPlaced.objects.filter(Delivered=True)

    context={
    'data2':data2,

    }
    return render(request,'admin/admin_payments.html',context)

def admin_dashboard(request):
    # total_orders = OrderPlaced.objects.count()

    # total_revenue = OrderPlaced.objects.aggregate(total_revenue=models.Sum('total_cost'))['total_revenue'] or 0
    # context={
    # 'total_orders':total_orders,
    # 'total_revenue':total_revenue,
    # }
    return render(request,'admin/admin_dashboard.html')




def add_customize_design(request):
    if request.method == "POST":
       
        
        image1 = request.FILES.get('image1')
        image2 = request.FILES.get('image2')
        image3 = request.FILES.get('image3')
        image4 = request.FILES.get('image4')
        image5 = request.FILES.get('image5')
       
        
        data=customized_Designed(image1=image1,image2=image2,image3=image3,image4=image4,image5=image5)
        data.save()
      # Fetch all products from the database
    return render(request, 'admin/add_customize_design.html')


def all_customer_info(request):
    all_customer = customer_user.objects.filter(is_customer=True)
    context={
        'all_customer':all_customer,
    }
    return render(request,'customer/customer_info.html',context)


def view_customerproducts(request,id):
   
    products = customer_user.objects.get(id=id)
    return render(request, 'customer/view_customer_products.html', {'product': products})



import sweetify
@login_required(login_url="/admin_login/")
def email_setting(request):
    if request.method == 'POST':
        email_form = EmailForm(request.POST)

        if email_form.is_valid():
            email_type = email_form.cleaned_data.get("email_type")
            smtp_username = email_form.cleaned_data.get("smtp_username")
            smtp_password = email_form.cleaned_data.get("smtp_password")
            smtp_server = email_form.cleaned_data.get("smtp_server")
            smtp_port = email_form.cleaned_data.get("smtp_port")
            smtp_security = email_form.cleaned_data.get("smtp_security")

            emailSetting.objects.update_or_create(email_type=email_type, defaults={
                'email_type': email_type, 'smtp_username': smtp_username, 'smtp_password': smtp_password, 'smtp_server': smtp_server, 'smtp_port': smtp_port, 'smtp_security': smtp_security
            })

            sweetify.success(request, title="Success", icon='success',
                             text='Email Setting Stored Successfully !!!')
            return redirect('/email-setting/')
        else:
            sweetify.error(request, title="error",
                           icon='danger', text='Failed !!!')
    else:
        email_form = EmailForm()
        if emailSetting.objects.all():
            email_data = emailSetting.objects.latest('id')
        else:
            email_data = None
    
    context = {
            'email_form': email_form,
            'email_data': email_data,
            "SettingsNavclass": "active",
            "settingsCollapseShow": "show",
            "EmailSettingsNavclass": "active"
    }

    return render(request, "email_setting.html", context)


from django.shortcuts import render, get_object_or_404
from .models import about_us  # Import your `about_us` model

def aboutus(request):
    # Retrieve the existing about_us object or create a new one if it doesn't exist
    about_us_obj, created = about_us.objects.get_or_create(pk=1)  # Assuming there's only one about_us object

    if request.method == 'POST':
        # Get the data from POST request
        history = request.POST.get('history')
        vision = request.POST.get('vision')
        mission = request.POST.get('mission')

        # Update the fields of the about_us object
        about_us_obj.history = history
        about_us_obj.vision = vision
        about_us_obj.mission = mission
        about_us_obj.save()

    # Pass the about_us object to the template for rendering
    return render(request, 'about_us.html', {'about_us_obj': about_us_obj})


def about_vr(request):
    data=about_us.objects.get()
    print('bbbbbbbbb',data)
    context={
        'about_us_obj':data,
    }
    return render(request,'about_vr.html',context)

def faq_questions(request):
    if request.method == 'POST':
        # Get the data from POST request
        question = request.POST.get('question')
        answers = request.POST.get('answers')

        data=faqs(question=question,answers=answers)
        data.save()

        return redirect('/faqs/')

    # Pass the about_us object to the template for rendering
    return render(request, 'faqs.html')

def display_faqs(request):
    data = faqs.objects.all()
    context={
    'data':data,
    }
    return render(request, 'display_faqs.html',context)



def add_events(request):
    if request.method == "POST":
        event_name = request.POST.get('event_name')
        event_date = request.POST.get('event_date')
        event_discount = request.POST.get('event_discount')
        Event(event_name=event_name, event_date=event_date, event_discount=event_discount).save()
    return render(request,'admin/events/add_events.html')

def view_events(request):
    events = Event.objects.all()
    context ={
        'events':events
    }
    return render(request,'admin/events/view_events.html',context)    

def edit_event(request,id):
    event = Event.objects.get(pk=id)

    if request.method == 'POST':
        Event.event_name = request.POST.get('event_name', '')
        Event.event_date = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        Event.event_discount = request.POST.get('event_discount', '')

        
        event.save()

        return redirect('/view_events/', id=event.id)

    return render(request, 'admin/events/edit_events.html', {'event': event})

def delete_event(request,id):
   event = Event.objects.filter(pk=id)
    
   if request.method == 'POST':
        event.delete()
        return redirect('/view_events/')


def inventory_listing(request):
    aggregated_products = product_seller.objects.values('title','brand', 'product_type', 'category','colour','size','selling_price').annotate(
        total_quantity=Sum('quantity')
        # total_ordered=Coalesce(Sum('quantity',output_field=IntegerField()), 0)
    ).annotate

    return render(request, 'admin/inventory_listing.html', {'products': aggregated_products})


def inventory_report(request):
    # Fetch all products from the database and annotate with total_quantity
    aggregated_products = product_seller.objects.values(
        'id', 'title', 'brand', 'product_type', 'category', 'colour', 'size', 'selling_price'
    ).annotate(
        total_quantity=Sum('quantity'),
    )

    # Calculate total ordered quantity for each product
    for product in aggregated_products:
        total_ordered = OrderPlaced.objects.filter(product_id=product['id']).aggregate(total_ordered=Sum('quantity'))['total_ordered'] or 0
        product['total_ordered'] = total_ordered

    # Calculate total returned quantity for each product
    for product in aggregated_products:
        total_returned = OrderPlaced.objects.filter(product_id=product['id'], returned=True).aggregate(total_returned=Sum('quantity'))['total_returned'] or 0
        product['total_returned'] = total_returned

    # Calculate total canceled quantity for each product
    for product in aggregated_products:
        total_canceled = CanceledOrder.objects.filter(product_id=product['id'], canceled=True).aggregate(total_canceled=Sum('quantity'))['total_canceled'] or 0
        product['total_canceled'] = total_canceled

    # Pass the fetched products to the template for rendering
    return render(request, 'admin/inventory_report.html', {'products': aggregated_products})


import os
def add_product(request):
    if request.method == 'POST':
        first_three_chars = 'PROD'
        remaining_digits = ''.join(random.choices( string.digits, k=6))
        prod_id = f'{first_three_chars}{remaining_digits}'

        # Assuming form fields are obtained from request.POST
        SKUId = request.POST.get('skuId')
        brand = request.POST.get('brand')
        product_id = request.POST.get('productId')
        size = request.POST.get('size')
        colour = request.POST.get('colour')
        title = request.POST.get('title')
        description = request.POST.get('description')
        product_type = request.POST.get('productType')
        category = request.POST.get('category')
        sub_category = request.POST.get('subCategory')
        actual_price = float(request.POST.get('actual_price', 0.0))
        discount_type = request.POST.get('discountType')
        discount_price = float(request.POST.get('discount_price', 0.0))  
        product_image = request.FILES.get('product_image')
        back_image = request.FILES.get('back_image')
        left_image = request.FILES.get('left_image')
        right_image = request.FILES.get('right_image')  
        video_file = request.FILES.get('product_video')

        # Save the video file to your desired location
        if video_file:
            with open(os.path.join(settings.MEDIA_ROOT, video_file.name), 'wb+') as destination:
                for chunk in video_file.chunks():
                    destination.write(chunk)

        quantity_str = request.POST.get('quantity')
        quantity = int(quantity_str) if quantity_str else 0
        gender = request.POST.get('gender')
        product_instruction = request.POST.get('productInstruction')
        manufacturer = request.POST.get('manufacturer')
        model_name = request.POST.get('modelName')
        model_number = request.POST.get('modelNumber')
        age_range_description = request.POST.get('ageRangeDescription')
        bullet_point = request.POST.get('bulletPoint')
        special_feature = request.POST.get('specialFeature')
        material_type = request.POST.get('materialType')
        manufacturer_contact = request.POST.get('manufacturerContact')
        product_story = request.POST.get('productStory')

        if discount_type == 'rupees':
            discount_price = float(request.POST.get('discount_price', 0.0))
            selling_price = actual_price - discount_price
            calculated_discount_percentage = (discount_price / actual_price) * 100
        elif discount_type == 'percentage':
            calculated_discount_percentage = float(request.POST.get('discount_price', 0.0))
            discount_price = (calculated_discount_percentage / 100) * actual_price
            selling_price = actual_price - discount_price
        else:
            # Handle other cases or set default values
            discount_price = 0.0
            calculated_discount_percentage = 0.0
            selling_price = actual_price

        new_product = product_seller(
                        prod_id=prod_id,

                        SKUId=SKUId,
                        brand=brand,
                        product_id=product_id,
                        size=size,
                        colour=colour,
                        title=title,
                        description=description,
                        product_type=product_type,
                        category=category,
                        sub_category=sub_category,
                        actual_price=actual_price,
                        discount_type=discount_type,
                        discount_price=discount_price,
                        product_image = product_image,
                        back_image = back_image,
                        left_image = left_image,
                        right_image = right_image,
                        product_video=video_file.name if video_file else None,

                        selling_price=selling_price,
                        quantity=quantity,
                        gender=gender,
                        product_instruction=product_instruction,
                        manufacturer=manufacturer,
                        model_name=model_name,
                        model_number=model_number,
                        age_range_description=age_range_description,
                        bullet_point=bullet_point,
                        special_feature=special_feature,
                        material_type=material_type,
                        manufacturer_contact=manufacturer_contact,
                        product_story=product_story,
                        discount_percentage=calculated_discount_percentage,
                    )

        new_product.save()
        print('dfghddddddddddddddd',new_product)

    return render(request, 'admin/admin_add_products.html')


def payment_setting(request):
    razorpay_settings = RazorpaySettings.objects.first()  # Retrieve settings
    
    if request.method == 'POST':
        if razorpay_settings is None:
            razorpay_settings = RazorpaySettings()  # Create a new instance if none exists
        
        razorpay_settings.razorpay_key_id = request.POST.get('razorpay_key_id')
        razorpay_settings.rezorpay_key_secret = request.POST.get('rezorpay_key_secret')
        razorpay_settings.save()  # Save the changes
        
        return redirect('payment_setting')  # Redirect back to the same page after saving
    
    context = {
        'razorpay_settings': razorpay_settings
    }
    return render(request, 'admin/payment_setting.html', context)

def collections_thought(request):
    # subcategories = ['BR_ambedkar_quotes', 'Chrapati_shivaji_maharaj_quotes', 'A_P_J_Abdul_kalam_tshirts', 'Bhagat_Singh', 'Budha_quotes', 'Birsa_munda', 'Savitribai_Phule', 'Martin_Luther_King', 'Sant_Kabir_Saheb', 'Sri_Guru_Nanak_Dev_ji', 'Osho_quotes']

    if request.method == "POST":
        subcategory = request.POST.get('subcategory')
        heading = request.POST.get('heading')
        heading1 = request.POST.get('heading1')
        heading2 = request.POST.get('heading2')
        heading3 = request.POST.get('heading3')
        heading4 = request.POST.get('heading4')
        heading5 = request.POST.get('heading5')
        heading6 = request.POST.get('heading6')
        paragraph1 = request.POST.get('paragraph1')
        paragraph2 = request.POST.get('paragraph2')
        paragraph3 = request.POST.get('paragraph3')
        paragraph4 = request.POST.get('paragraph4')
        paragraph5 = request.POST.get('paragraph5')
        paragraph6 = request.POST.get('paragraph6')
        paragraph7 = request.POST.get('paragraph7')
        

        # Create a new Collections_Thought object and save it
        new_thought = Collections_Thought(
            subcategory=subcategory, 
            heading=heading,
            heading1=heading1,
            heading2=heading2,
            heading3=heading3,
            heading4=heading4,
            heading5=heading5,
            heading6=heading6,
            paragraph1=paragraph1,
            paragraph2=paragraph2,
            paragraph3=paragraph3,
            paragraph4=paragraph4,
            paragraph5=paragraph5,
            paragraph6=paragraph6,
            paragraph7=paragraph7
        )
        new_thought.save()   

    return render(request,'admin/collections_thought.html')


def edit_collections_thought(request):
    thoughts = Collections_Thought.objects.all()  
    context = {
        'thoughts': thoughts 
    }       

    return render(request,'admin/edit_collections_thought.html',context)

def edit_collections(request,id):
    thoughts = Collections_Thought.objects.get(id=id) 
    if request.method == "POST":
        thoughts.subcategory = request.POST.get('subcategory')
        thoughts.heading = request.POST.get('heading')       
        thoughts.heading1 = request.POST.get('heading1')
        thoughts.heading2 = request.POST.get('heading2')
        thoughts.heading3 = request.POST.get('heading3')
        thoughts.heading4 = request.POST.get('heading4')
        thoughts.heading5 = request.POST.get('heading5')
        thoughts.heading6 = request.POST.get('heading6')
        thoughts.paragraph1 = request.POST.get('paragraph1')
        thoughts.paragraph2 = request.POST.get('paragraph2')
        thoughts.paragraph3 = request.POST.get('paragraph3')
        thoughts.paragraph4 = request.POST.get('paragraph4')
        thoughts.paragraph5 = request.POST.get('paragraph5')
        thoughts.paragraph6 = request.POST.get('paragraph6')
        thoughts.paragraph7 = request.POST.get('paragraph7')
        thoughts.save()
    context = {
        'thoughts': thoughts 
    }       
    return render(request,'admin/edit_collections.html',context)

def delete_collections(request,id):
    thoughts = Collections_Thought.objects.get(id=id)   
       
    thoughts.delete()
    return redirect('edit_collections_thought')